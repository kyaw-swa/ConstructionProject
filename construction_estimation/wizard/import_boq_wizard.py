import base64
import io
import logging

import openpyxl

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Column indices (0-based) – Detail of Measurements sheet
#   A  B           C    D  E       F      G      H      I      J      K      L        M
#   No Particular  No.  x  Factor  L(ft)  L(in)  B(ft)  B(in)  H(ft)  H(in)  Content  Unit
COL_ITEM_NO   = 0
COL_PARTICULAR = 1
COL_NO        = 2
COL_X         = 3
COL_FACTOR    = 4
COL_L_FT      = 5
COL_L_IN      = 6
COL_B_FT      = 7
COL_B_IN      = 8
COL_H_FT      = 9
COL_H_IN      = 10
COL_CONTENT   = 11
COL_UNIT      = 12


class ImportBOQWizard(models.TransientModel):
    _name = 'construction.import.boq.wizard'
    _description = 'Import BOQ from Excel'

    estimate_id = fields.Many2one(
        'construction.project.estimate',
        string='Project Estimate',
        required=True,
    )
    excel_file = fields.Binary(string='BOQ Excel File', attachment=True)
    filename = fields.Char()
    sheet_name = fields.Char(
        placeholder='Leave empty to use the first sheet.',
    )
    header_rows = fields.Integer(
        default=6,
        help=(
            'Rows to skip at the top (default 6: Name of Work, '
            'Name of Builder, Detail of Measurements, column header, '
            'L/B/H labels, Ft/In labels).'
        ),
    )
    result_summary = fields.Text(readonly=True)

    # ── Public action ────────────────────────────────────────────────────────

    def action_import(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))

        rows = self._read_excel_rows()
        stats = self._process_rows(rows)
        self.result_summary = self._build_summary(stats)

        # Re-open the same wizard record so the result summary becomes visible.
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    # ── Excel reading ────────────────────────────────────────────────────────

    def _read_excel_rows(self):
        file_data = base64.b64decode(self.excel_file)
        wb = openpyxl.load_workbook(
            io.BytesIO(file_data), read_only=True, data_only=True,
        )
        if self.sheet_name:
            if self.sheet_name not in wb.sheetnames:
                raise UserError(
                    _("Sheet '%s' not found. Available sheets: %s")
                    % (self.sheet_name, ', '.join(wb.sheetnames))
                )
            ws = wb[self.sheet_name]
        else:
            ws = wb.active

        skip = self.header_rows or 0
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < skip:
                continue
            rows.append(row)
        return rows

    # ── Row processing ───────────────────────────────────────────────────────

    def _process_rows(self, rows):
        AcModel     = self.env['construction.ac']
        Measurement = self.env['construction.measurement']
        EstimateLine = self.env['construction.estimate.line']

        stats = {
            'lines': 0,
            'ac_created': 0,
            'ac_existing': 0,
            'meas_created': 0,
            'meas_existing': 0,
            'skipped': 0,
            'errors': [],
        }
        ac_cache   = {}
        meas_cache = {}

        current_ac        = None
        current_unit      = ''
        current_meas_name = None

        for row_idx, row in enumerate(rows, start=(self.header_rows or 0) + 1):
            try:
                item_no   = self._cell(row, COL_ITEM_NO)
                particular = self._cell(row, COL_PARTICULAR)
                unit_col  = self._cell(row, COL_UNIT)

                if not particular and not item_no:
                    stats['skipped'] += 1
                    continue

                is_data = self._is_data_row(row)

                if is_data:
                    if not current_ac:
                        stats['skipped'] += 1
                        continue

                    quantity = self._get_quantity(row)
                    if not quantity:
                        stats['skipped'] += 1
                        continue

                    meas_name = current_meas_name or (
                        str(particular).strip() if particular else ''
                    )
                    measurement = None
                    if meas_name:
                        measurement = self._get_or_create_measurement(
                            Measurement, meas_cache, current_ac, meas_name, stats,
                        )

                    EstimateLine.create({
                        'estimate_id':    self.estimate_id.id,
                        'ac_id':          current_ac.id,
                        'measurement_id': measurement.id if measurement else False,
                        'quantity':       quantity,
                        'unit':           current_unit,
                    })
                    stats['lines'] += 1

                elif item_no and particular:
                    # New A/C section header
                    current_ac = self._get_or_create_ac(
                        AcModel, ac_cache, str(particular).strip(), stats,
                    )
                    current_unit = str(unit_col).strip() if unit_col else ''
                    current_meas_name = None

                elif particular and current_ac:
                    # Measurement sub-item within the current A/C
                    current_meas_name = str(particular).strip()

                else:
                    stats['skipped'] += 1

            except Exception as exc:
                msg = f'Row {row_idx}: {exc}'
                _logger.warning('BOQ import – %s', msg)
                stats['errors'].append(msg)

        return stats

    # ── Row-type helpers ─────────────────────────────────────────────────────

    def _is_data_row(self, row):
        """True when the row carries actual measurement/quantity data."""
        c_val = self._cell(row, COL_NO)
        d_val = self._cell(row, COL_X)
        for val in (c_val, d_val):
            if val is None:
                continue
            try:
                if float(val):
                    return True
            except (TypeError, ValueError):
                pass
        return False

    def _get_quantity(self, row):
        """Return quantity: prefer the pre-computed Content cell; fall back to L×B×H."""
        content = self._float(row, COL_CONTENT)
        if content:
            return content

        no     = self._float(row, COL_NO)     or 1.0
        x      = self._float(row, COL_X)      or 1.0
        factor = self._float(row, COL_FACTOR) or 1.0

        l = self._float(row, COL_L_FT) + self._float(row, COL_L_IN) / 12.0
        b = self._float(row, COL_B_FT) + self._float(row, COL_B_IN) / 12.0
        h = self._float(row, COL_H_FT) + self._float(row, COL_H_IN) / 12.0

        dims = [d for d in (l, b, h) if d]
        quantity = no * x * factor
        for d in dims:
            quantity *= d
        return quantity

    # ── A/C and Measurement cache helpers ───────────────────────────────────

    def _get_or_create_ac(self, Model, cache, name, stats):
        if name in cache:
            stats['ac_existing'] += 1
            return cache[name]
        rec = Model.search([('name', '=', name)], limit=1)
        if rec:
            stats['ac_existing'] += 1
        else:
            rec = Model.create({'name': name})
            stats['ac_created'] += 1
        cache[name] = rec
        return rec

    def _get_or_create_measurement(self, Model, cache, ac, name, stats):
        key = (ac.id, name)
        if key in cache:
            stats['meas_existing'] += 1
            return cache[key]
        rec = Model.search([('ac_id', '=', ac.id), ('name', '=', name)], limit=1)
        if rec:
            stats['meas_existing'] += 1
        else:
            rec = Model.create({'ac_id': ac.id, 'name': name})
            stats['meas_created'] += 1
        cache[key] = rec
        return rec

    # ── Low-level cell utilities ─────────────────────────────────────────────

    @staticmethod
    def _cell(row, col):
        if not row or col >= len(row):
            return None
        val = row[col]
        return val if val not in (None, '') else None

    @staticmethod
    def _float(row, col):
        if not row or col >= len(row):
            return 0.0
        val = row[col]
        if val is None or val == '':
            return 0.0
        try:
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    # ── Summary builder ──────────────────────────────────────────────────────

    @staticmethod
    def _build_summary(stats):
        lines = [
            f"Estimation lines created : {stats['lines']}",
            f"A/C records created      : {stats['ac_created']}",
            f"A/C records reused       : {stats['ac_existing']}",
            f"Measurements created     : {stats['meas_created']}",
            f"Measurements reused      : {stats['meas_existing']}",
            f"Rows skipped             : {stats['skipped']}",
        ]
        if stats['errors']:
            lines.append(f"\nWarnings ({len(stats['errors'])}):")
            lines.extend(f"  {e}" for e in stats['errors'][:10])
            if len(stats['errors']) > 10:
                lines.append(f"  … and {len(stats['errors']) - 10} more")
        return '\n'.join(lines)
